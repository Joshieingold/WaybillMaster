import openpyxl
import pyautogui
import time
from tkinter import filedialog
import clipboard
import pandas as pd

tech_lst = ["B176", "B315", "NB00", "NB02", "NB03", "NB04", "NB05", "NB06", "NB07", "NB08", "NB09", "NE15", "NE23", "NE26", "NE74", "NF03", "NF04", "NF21", "NF23", "NF24", "NF26", "NF27", "NF34", "NF35", "NF36", "NF38", "NF41", "NF44", "NF45", "NF56", "NF69", "Nm43", "NN17", "NN19", "NN24", "NN46", "NN51", "NN53", "NN54", "NN57", "NN58", "NN60", "NN61", "NN62", "NN63", "NN64", "NN66", "NN67", "NN68", "NN71", "NN72", "NN73", "NN75", "NN79", "NN80", "NS07", "NS09", "NS15", "NS17", "NS22", "NS23", "NS32", "NS34", "NS36", "NS41", "NS53", "NS55", "NS57", "NS68", "ns71", "NS73", "Ns77", "NS83", "NS84", "NS85", "NS86", "NS90", "NS91", "NS92", "NS93", "NS94", "NS98"]
date = "Qty - May 22"
num_down = 16

def open_file_dialog():
    print("[INFO] Select the Excel file to process...")
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        print("[WARNING] No file selected.")
        return

    print(f"[INFO] Processing file: {file_path}")
    result_list = process_file(file_path)
    if result_list:
        copy_data_to_excel(result_list)

def process_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        return process_sheet(sheet)
    except Exception as e:
        print(f"[ERROR] Failed to process file: {e}")
        return []

def process_sheet(sheet):
    result_lst = []
    devices = [
        "XB8", "XB7", "XI6", "XIONE", "PODS", "ONTS",
        "SCHB1AEW", "SCHC2AEW", "SCHC3AEW", "SCXI11BEI-ENTOS",
        "WNXB11ABR", "WNXL11BWL", "NOVA2002"
    ]
    device_mapping = {
        "CGM4981COM": "XB8",
        "CGM4331COM": "XB7", "TG4482A": "XB7",
        "IPTVARXI6HD": "XI6", "IPTVTCXI6HD": "XI6",
        "SCXI11BEI": "XIONE",
        "XE2SGROG1": "PODS",
        "XS010XB": "ONTS", "XS010XQ": "ONTS", "XS020XONT": "ONTS",
        "SCHB1AEW": "SCHB1AEW",
        "SCHC2AEW": "SCHC2AEW",
        "SCHC3AE0": "SCHC3AEW",
        "SCXI11BEI-ENTOS": "SCXI11BEI-ENTOS",
        "WNXB11ABR": "WNXB11ABR",
        "WNXL11BWL": "WNXL11BWL",
        "NOVA2002": "NOVA2002",
    }

    def update_totals(totals, item_code, allowed_devices):
        if item_code in device_mapping:
            device = device_mapping[item_code]
            if device in allowed_devices:
                totals[device] += 1

    for tech in tech_lst:
        print(f"[PROCESSING] Technician: {tech}")
        tech_totals = {device: 0 for device in devices}
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            contractor_id = str(row[7].value).strip() if row[7].value else ""
            item_code = row[5].value
            inventory_type = str(row[9].value).strip() if row[9].value else ""

            if contractor_id.upper() == tech.upper() and inventory_type == f"TEC.Subready.{tech}":
                update_totals(tech_totals, item_code, devices)

        formatted = format_totals(tech_totals, devices)
        result_lst.append(formatted)

    print(f"[INFO] Finished processing {len(result_lst)} technicians.")
    return result_lst

def format_totals(totals, device_order):
    return '\n'.join(str(totals[device]) for device in device_order)

def copy_data_to_excel(data_lst):
    print("[INFO] Starting auto copy to Excel in 5 seconds...")
    time.sleep(5)
    for index, data in enumerate(data_lst):
        print(f"[COPY] Technician {index + 1}/{len(data_lst)}: Pasting data...")
        pyautogui.write(date)
        pyautogui.press("down")
        clipboard.copy(data)
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'v')
        for i in range(num_down):
            pyautogui.press("down")
    print("[DONE] All data copied to Excel.")

# Run the tool
open_file_dialog()
